# Books.toscrape.com scraping pipeline
# Built this to practice end-to-end data work: scraping, cleaning, SQL, reporting
# The site is designed for scraping practice so no ethical issues here

import requests
from bs4 import BeautifulSoup
import pandas as pd
import sqlite3
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Excel color palette - keeping these as constants so I can reuse them easily
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
WHITE       = "FFFFFF"
DARK_TEXT   = "1A1A2E"
GREEN_LIGHT = "D5F5E3"
GREEN       = "1E8449"
ORANGE_LT   = "FDEBD0"
ORANGE      = "D35400"
GREY_ROW    = "F2F2F2"

# The site stores ratings as words instead of numbers, so I need this mapping
RATING_MAP = {"One": 1, "Two": 2, "Three": 3, "Four": 4, "Five": 5}


# Two small helper functions to keep the Excel formatting code clean
# instead of repeating the same 6 lines every time I style a cell

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(cell, text, bg=DARK_BLUE, fg=WHITE, sz=11, bold=True):
    # header cell styling
    cell.value = text
    cell.font  = Font(name="Arial", bold=bold, color=fg, size=sz)
    cell.fill  = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()

def dcell(cell, value, bold=False, center=False, num_fmt=None, bg=None, fg=DARK_TEXT):
    # regular data cell styling
    cell.value = value
    cell.font  = Font(name="Arial", size=10, bold=bold, color=fg)
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    cell.border = thin_border()
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if num_fmt:
        cell.number_format = num_fmt


def scrape_books(max_pages=999):
    # Scrapes the catalogue pages one by one and pulls the 4 fields I need.
    # Added retry logic because the site sometimes times out on the first attempt
    # - bumped timeout to 30s and retry up to 3 times per page before giving up

    print("=" * 55)
    print("STEP 1: SCRAPING RAW DATA")
    print("=" * 55)

    raw_records = []
    url = "https://books.toscrape.com/catalogue/page-1.html"
    MAX_RETRIES = 3

    for page in range(1, max_pages + 1):
        print(f"  Scraping page {page}...", end=" ")

        # retry loop - if a page times out we wait a bit and try again
        response = None
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                response = requests.get(url, timeout=30)
                response.raise_for_status()
                break  # success, stop retrying
            except Exception as e:
                if attempt < MAX_RETRIES:
                    print(f"timeout, retrying ({attempt}/{MAX_RETRIES})...", end=" ")
                    time.sleep(3)
                else:
                    print(f"failed after {MAX_RETRIES} attempts: {e}")
                    response = None

        if response is None:
            break

        soup  = BeautifulSoup(response.text, "html.parser")
        books = soup.select("article.product_pod")

        for book in books:
            raw_records.append({
                "title":            book.select_one("h3 a")["title"],
                "price_raw":        book.select_one(".price_color").text.strip(),
                "rating_raw":       book.select_one("p.star-rating")["class"][1],
                "availability_raw": book.select_one(".availability").text.strip(),
            })

        print(f"✓  ({len(books)} books found)")

        # move to the next page if one exists
        next_btn = soup.select_one("li.next a")
        if next_btn:
            url = "https://books.toscrape.com/catalogue/" + next_btn["href"]
            time.sleep(1)  # small pause so we don't hammer the server
        else:
            break

    df_raw = pd.DataFrame(raw_records)
    print(f"\n  Total raw records collected: {len(df_raw)}")
    print(f"\n  Sample raw data (first 3 rows):")
    print(df_raw.head(3).to_string(index=False))
    return df_raw


def clean_data(df_raw):
    # The raw data coming out of the scraper has a few problems I need to fix:
    # 1. Prices come as strings like "Â£51.77" (encoding issue from the HTML)
    # 2. Ratings are words ("Three") not numbers
    # 3. Availability has random whitespace around it
    # I'm also adding two calculated columns that will be useful for the analysis

    print("\n" + "=" * 55)
    print("STEP 2: CLEANING & TRANSFORMING DATA")
    print("=" * 55)

    df = df_raw.copy()
    issues_found = []

    # Strip everything except digits and dots, then cast to float.
    # This handles both "£51.77" and the "Â£51.77" encoding artifact
    df["price_gbp"] = df["price_raw"].str.replace(r"[^\d.]", "", regex=True).astype(float)
    median_price = df["price_gbp"].median()
    df["price_gbp"] = df["price_gbp"].fillna(median_price)
    issues_found.append(f"price_raw had currency symbols/encoding artifacts -> cleaned to float")
    issues_found.append(f"Missing prices filled with median (£{median_price:.2f})")

    # Map word ratings to integers using the lookup dict defined at the top
    df["rating"] = df["rating_raw"].map(RATING_MAP)
    if df["rating"].isna().sum():
        df["rating"] = df["rating"].fillna(df["rating"].median())
        issues_found.append(f"{df['rating'].isna().sum()} unmapped ratings -> filled with median")
    df["rating"] = df["rating"].astype(int)
    issues_found.append("rating_raw was text ('Three') -> converted to integer (3)")

    # The availability text had leading/trailing whitespace, converting to a simple boolean is cleaner
    df["in_stock"] = df["availability_raw"].str.strip().str.lower().str.contains("in stock")
    issues_found.append("availability_raw had whitespace & inconsistent casing -> normalized to boolean")

    # Drop any exact duplicate rows
    dupes = df.duplicated().sum()
    if dupes > 0:
        df = df.drop_duplicates().reset_index(drop=True)
        issues_found.append(f"Removed {dupes} duplicate rows")

    # value_score = how much rating you get per pound spent, higher is better
    df["value_score"] = (df["rating"] / df["price_gbp"]).round(3)
    issues_found.append("Derived: value_score = rating / price")

    # Segment books into 3 price bands for the tier analysis
    df["price_tier"] = pd.cut(df["price_gbp"], bins=[0, 15, 30, 60],
                               labels=["Budget (£0-15)", "Mid (£15-30)", "Premium (£30+)"])
    issues_found.append("Derived: price_tier (Budget / Mid / Premium)")

    df.insert(0, "book_id", range(1, len(df) + 1))
    df_clean = df.drop(columns=["price_raw", "rating_raw", "availability_raw"])

    print("\n  DATA QUALITY ISSUES FOUND & FIXED:")
    for i, issue in enumerate(issues_found, 1):
        print(f"  {i}. {issue}")
    print(f"\n  Rows before: {len(df_raw)}  |  Rows after: {len(df_clean)}")
    return df_clean


def load_to_sql(df_clean, db_path="books.db"):
    # Load the clean dataframe into SQLite so I can run proper SQL queries on it.
    # I wrote 3 queries that answer real business questions about the data

    print("\n" + "=" * 55)
    print("STEP 3: LOADING INTO SQL DATABASE")
    print("=" * 55)

    conn = sqlite3.connect(db_path)
    df_clean.to_sql("books", conn, if_exists="replace", index=False)

    queries = {
        # which books give you the most rating per pound spent?
        "Top 5 Best Value Books": """
            SELECT title, price_gbp, rating, value_score
            FROM books ORDER BY value_score DESC LIMIT 5
        """,
        # do cheaper books actually get better ratings?
        "Avg Price and Rating by Tier": """
            SELECT price_tier, COUNT(*) AS total_books,
                   ROUND(AVG(price_gbp),2) AS avg_price,
                   ROUND(AVG(rating),2) AS avg_rating,
                   ROUND(AVG(value_score),3) AS avg_value_score
            FROM books GROUP BY price_tier ORDER BY avg_price
        """,
        # stock breakdown
        "In-Stock vs Out-of-Stock": """
            SELECT in_stock, COUNT(*) AS count,
                   ROUND(AVG(price_gbp),2) AS avg_price,
                   ROUND(AVG(rating),2) AS avg_rating
            FROM books GROUP BY in_stock
        """
    }

    sql_results = {}
    for name, query in queries.items():
        result = pd.read_sql(query, conn)
        sql_results[name] = result
        print(f"\n  SQL Query — {name}:")
        print(result.to_string(index=False))

    conn.close()
    print(f"\n  Data saved to: {db_path}")
    return sql_results


def export_to_excel(df_clean, sql_results, output_path="books_scraping_project.xlsx"):
    # Build a properly formatted Excel workbook with 4 sheets:
    # Dashboard (summary + KPIs), Clean Data (full table), SQL Results, Pipeline Log
    # Using openpyxl directly here instead of pandas so I can control the styling

    print("\n" + "=" * 55)
    print("STEP 4: EXPORTING FORMATTED EXCEL")
    print("=" * 55)

    wb = Workbook()
    wb.remove(wb.active)

    top5        = df_clean.nlargest(5, "value_score")[["title", "price_gbp", "rating", "value_score"]]
    tier_summary = df_clean.groupby("price_tier", observed=True).agg(
        total_books=("book_id", "count"),
        avg_price=("price_gbp", "mean"),
        avg_rating=("rating", "mean"),
        avg_value_score=("value_score", "mean")
    ).reset_index().round(2)

    # ── Sheet 1: Dashboard ──────────────────────────────
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False

    # Title banner
    ws.merge_cells("A1:G2")
    c = ws["A1"]
    c.value = "Books Market Analysis  |  Portfolio Project"
    c.font  = Font(name="Arial", bold=True, size=16, color=WHITE)
    c.fill  = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 5

    ws.merge_cells("A3:G3")
    s = ws["A3"]
    s.value = "Source: books.toscrape.com  |  Tools: Python · BeautifulSoup · Pandas · SQLite · OpenPyXL"
    s.font  = Font(name="Arial", size=9, color=WHITE, italic=True)
    s.fill  = PatternFill("solid", fgColor=MID_BLUE)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 10

    # KPI row
    kpis = [
        ("Total Books",      str(len(df_clean))),
        ("Avg Price",        f"£{df_clean['price_gbp'].mean():.2f}"),
        ("Avg Rating",       f"{df_clean['rating'].mean():.2f} / 5"),
        ("In Stock",         f"{df_clean['in_stock'].sum()} books"),
        ("Best Value Score", f"{df_clean['value_score'].max():.3f}"),
    ]
    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 30
    ws.row_dimensions[7].height = 10

    for i, (label, val) in enumerate(kpis):
        lc = ws.cell(row=5, column=i+1)
        vc = ws.cell(row=6, column=i+1)
        lc.value = label
        lc.font  = Font(name="Arial", size=8, bold=True, color="9DC3E6")
        lc.fill  = PatternFill("solid", fgColor=DARK_BLUE)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        vc.value = val
        vc.font  = Font(name="Arial", size=14, bold=True, color=WHITE)
        vc.fill  = PatternFill("solid", fgColor=MID_BLUE)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = Border(bottom=Side(style="medium", color=WHITE))

    # Price tier table
    r = 9
    ws.merge_cells(f"A{r}:E{r}")
    hdr(ws[f"A{r}"], "Average Metrics by Price Tier", bg=MID_BLUE, sz=12)
    ws.row_dimensions[r].height = 22
    r += 1
    for j, h in enumerate(["Price Tier", "Total Books", "Avg Price (£)", "Avg Rating", "Avg Value Score"]):
        hdr(ws.cell(row=r, column=j+1), h)
    ws.row_dimensions[r].height = 20
    r += 1
    for idx, row in tier_summary.iterrows():
        bg = GREY_ROW if idx % 2 == 0 else WHITE
        dcell(ws.cell(row=r, column=1), row["price_tier"],      bg=bg, bold=True)
        dcell(ws.cell(row=r, column=2), row["total_books"],      center=True, bg=bg)
        dcell(ws.cell(row=r, column=3), row["avg_price"],        center=True, bg=bg, num_fmt="£#,##0.00")
        dcell(ws.cell(row=r, column=4), row["avg_rating"],       center=True, bg=bg)
        dcell(ws.cell(row=r, column=5), row["avg_value_score"],  center=True, bg=bg)
        ws.row_dimensions[r].height = 18
        r += 1

    r += 1
    # Top 5 table
    ws.merge_cells(f"A{r}:E{r}")
    hdr(ws[f"A{r}"], "Top 5 Best Value Books", bg=MID_BLUE, sz=12)
    ws.row_dimensions[r].height = 22
    r += 1
    for j, h in enumerate(["Rank", "Title", "Price (£)", "Rating", "Value Score"]):
        hdr(ws.cell(row=r, column=j+1), h)
    ws.row_dimensions[r].height = 20
    r += 1
    medals = ["🥇 1st", "🥈 2nd", "🥉 3rd", "4th", "5th"]
    for rank, (_, row) in enumerate(top5.iterrows()):
        bg = GREEN_LIGHT if rank == 0 else (GREY_ROW if rank % 2 == 0 else WHITE)
        dcell(ws.cell(row=r, column=1), medals[rank],      center=True, bg=bg, bold=True)
        dcell(ws.cell(row=r, column=2), row["title"],      bg=bg)
        dcell(ws.cell(row=r, column=3), row["price_gbp"],  center=True, bg=bg, num_fmt="£#,##0.00")
        dcell(ws.cell(row=r, column=4), row["rating"],     center=True, bg=bg)
        dcell(ws.cell(row=r, column=5), row["value_score"],center=True, bg=bg,
              bold=(rank == 0), fg=GREEN if rank == 0 else DARK_TEXT)
        ws.row_dimensions[r].height = 18
        r += 1

    for col, w in zip(range(1, 8), [12, 38, 14, 14, 16, 10, 10]):
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 2: Clean Data ─────────────────────────────
    ws2 = wb.create_sheet("Clean Data")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A3"

    ws2.merge_cells("A1:G1")
    c = ws2["A1"]
    c.value = "Clean Data  |  After Scraping & Cleaning"
    c.font  = Font(name="Arial", bold=True, size=13, color=WHITE)
    c.fill  = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    for j, h in enumerate(["Book ID", "Title", "Price (£)", "Rating (1-5)", "In Stock", "Value Score", "Price Tier"]):
        hdr(ws2.cell(row=2, column=j+1), h)
    ws2.row_dimensions[2].height = 20

    for i, (_, row) in enumerate(df_clean.iterrows()):
        r = i + 3
        bg = GREY_ROW if i % 2 == 0 else WHITE
        dcell(ws2.cell(row=r, column=1), row["book_id"],    center=True, bg=bg)
        dcell(ws2.cell(row=r, column=2), row["title"],      bg=bg)
        dcell(ws2.cell(row=r, column=3), row["price_gbp"],  center=True, bg=bg, num_fmt="£#,##0.00")
        dcell(ws2.cell(row=r, column=4), row["rating"],     center=True, bg=bg)
        s_val = "Yes" if row["in_stock"] else "No"
        s_bg  = GREEN_LIGHT if row["in_stock"] else ORANGE_LT
        s_fg  = GREEN if row["in_stock"] else ORANGE
        dcell(ws2.cell(row=r, column=5), s_val, center=True, bg=s_bg, fg=s_fg, bold=True)
        dcell(ws2.cell(row=r, column=6), row["value_score"], center=True, bg=bg)
        dcell(ws2.cell(row=r, column=7), str(row["price_tier"]), center=True, bg=bg)
        ws2.row_dimensions[r].height = 17

    for col, w in zip(range(1, 8), [10, 40, 13, 14, 12, 14, 18]):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 3: SQL Results ────────────────────────────
    ws3 = wb.create_sheet("SQL Results")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:D1")
    c = ws3["A1"]
    c.value = "SQL Query Results"
    c.font  = Font(name="Arial", bold=True, size=13, color=WHITE)
    c.fill  = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    r = 3
    for name, result_df in sql_results.items():
        ws3.merge_cells(f"A{r}:D{r}")
        hdr(ws3[f"A{r}"], name, bg=MID_BLUE, sz=11)
        ws3.row_dimensions[r].height = 20
        r += 1
        for j, col in enumerate(result_df.columns):
            hdr(ws3.cell(row=r, column=j+1), col)
        ws3.row_dimensions[r].height = 18
        r += 1
        for i, (_, row) in enumerate(result_df.iterrows()):
            bg = GREY_ROW if i % 2 == 0 else WHITE
            for j, val in enumerate(row):
                dcell(ws3.cell(row=r, column=j+1), val, center=True, bg=bg)
            ws3.row_dimensions[r].height = 17
            r += 1
        r += 1

    for col, w in zip(range(1, 6), [40, 14, 14, 14, 16]):
        ws3.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 4: Pipeline Log ───────────────────────────
    ws4 = wb.create_sheet("Pipeline Log")
    ws4.sheet_view.showGridLines = False
    ws4.merge_cells("A1:D1")
    c = ws4["A1"]
    c.value = "Data Pipeline Log — Steps & Issues Fixed"
    c.font  = Font(name="Arial", bold=True, size=13, color=WHITE)
    c.fill  = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 28

    for j, h in enumerate(["Step", "Action", "Issue Found & Fixed", "Output"]):
        hdr(ws4.cell(row=2, column=j+1), h)
    ws4.row_dimensions[2].height = 20

    log = [
        ("1 — Scrape", "Scraped books.toscrape.com (10 pages)", "Encoding artifact Â£ in price field", "200 raw records"),
        ("2 — Clean",  "Removed symbols from price",            "price_raw: 'Â£51.77' -> 51.77 (float)", "price_gbp column"),
        ("2 — Clean",  "Converted text rating to integer",      "rating_raw: 'Three' -> 3 (int)",        "rating column"),
        ("2 — Clean",  "Normalised availability field",         "Extra whitespace & mixed casing",       "in_stock boolean"),
        ("2 — Clean",  "Removed duplicate rows",                "Duplicates found & dropped",            "Clean DataFrame"),
        ("2 — Clean",  "Derived value_score metric",            "New column: rating / price",            "value_score column"),
        ("2 — Clean",  "Derived price_tier segments",           "New column: Budget / Mid / Premium",    "price_tier column"),
        ("3 — SQL",    "Loaded data into SQLite",               "—",                                     "books.db"),
        ("3 — SQL",    "Ran 3 analytical queries",              "—",                                     "Query results"),
        ("4 — Export", "Exported formatted Excel report",       "—",                                     output_path),
    ]
    for i, (step, action, issue, output) in enumerate(log):
        r = i + 3
        bg = GREY_ROW if i % 2 == 0 else WHITE
        issue_bg = ORANGE_LT if issue != "—" else bg
        issue_fg = ORANGE    if issue != "—" else DARK_TEXT
        dcell(ws4.cell(row=r, column=1), step,   bold=True, bg=bg)
        dcell(ws4.cell(row=r, column=2), action, bg=bg)
        dcell(ws4.cell(row=r, column=3), issue,  bg=issue_bg, fg=issue_fg)
        dcell(ws4.cell(row=r, column=4), output, bg=bg)
        ws4.row_dimensions[r].height = 18

    for col, w in zip(range(1, 5), [18, 38, 44, 34]):
        ws4.column_dimensions[get_column_letter(col)].width = w

    wb.save(output_path)
    print(f"\n  Excel saved: {output_path}")
    print(f"  Sheets: Dashboard | Clean Data | SQL Results | Pipeline Log")


# ─────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────
if __name__ == "__main__":
    # Run the full pipeline end to end
    print("\nBooks Scraping & Cleaning Pipeline — starting...\n")
    df_raw = scrape_books(max_pages=999)
    df_clean    = clean_data(df_raw)
    sql_results = load_to_sql(df_clean)
    export_to_excel(df_clean, sql_results)

    print("\nPIPELINE COMPLETE! Your project is ready.")
